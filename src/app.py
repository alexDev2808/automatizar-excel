import pandas as pd
from tabulate import tabulate
from tablas_sql import *
from openpyxl import load_workbook
from datetime import date

wb = load_workbook('PNuevoIngreso.xlsx')
hoja = wb.active

sin_datos = []

def generar_formato_tabla():
    datos = []
    headers = ["NoEmp","Nombre","APP","APM","Puesto","Empresa","Fecha","Correo"]
    formato_datos = []

    for fila in hoja.iter_rows(values_only=True):
        datos.append(dict(zip(headers, fila)))

    for index, usuario in enumerate(datos):
        if index == 0:
            continue
        
        pre_formato = dict()
        
        if usuario["Empresa"] == "BANCOR":
            pre_formato["id_area"] = 6
        else:
            pre_formato["id_area"] = 3
            
        pre_formato["id_empleado"] = f"0{str(usuario["NoEmp"])}"
        pre_formato["id_funcion"] = buscar_en_tabla(puestos, "puesto", usuario["Puesto"])
        pre_formato["app"] = usuario["APP"]
        pre_formato["apm"] = usuario["APM"]
        pre_formato["nombre"] = usuario["Nombre"]
        pre_formato["activo"] = 1
        pre_formato["pass"] = f"{usuario["APP"]}0{str(usuario["NoEmp"])}"
        pre_formato["id_area_res"] = buscar_colaboradores(responsables_area, "res_area", usuario["Puesto"])
        pre_formato["tc"] = buscar_colaboradores(nivel_puesto, "nivelPuesto", usuario["Puesto"])
        pre_formato["mail"] = usuario["Correo"]
        pre_formato["id_areat"] = buscar_colaboradores(areas_internas, "areat", usuario["Puesto"])
        pre_formato["id_area_res2"] = 5
        pre_formato["id_area_res3"] = ""
        pre_formato["perm_fsm"] = ""
        pre_formato["tipoPuesto"] = buscar_colaboradores(tipo_puesto, "tipoPuesto", usuario["Puesto"])
        
        formato_datos.append(pre_formato)

    return formato_datos


def buscar_en_tabla(tabla, nombre_columna, valor_buscar):
    for item in tabla:
        if valor_buscar == item[f"nombre_{nombre_columna}"]:
            return item[f"id_{nombre_columna}"]
    sin_datos.append(dict( columna = f"id_{nombre_columna}", valor_buscado = valor_buscar ))
    return None

def buscar_colaboradores(tabla, nombre_columna, valor_buscar):
    for item in tabla:
        for colaborador in item["colaboradores"]:
            if colaborador == valor_buscar:
                return item[f"id_{nombre_columna}"]
    sin_datos.append(dict( columna = f"Colaborador: ", valor_buscado = valor_buscar ))
    return None

def generar_tabla_excel(datos):
    id_empleado = []
    id_funcion = []
    id_area = []
    app = []
    apm = []
    nombre = []
    activo = []
    passw = []
    id_area_res = []
    tc = []
    mail = []
    id_areat = []
    id_area_res2 = []
    id_area_res3 = []
    perm_fsm = []
    tipoPuesto = []
    
    for fila in datos:
        id_empleado.append(fila["id_empleado"])
        id_funcion.append(fila["id_funcion"])
        id_area.append(fila["id_area"])
        app.append(fila["app"])
        apm.append(fila["apm"])
        nombre.append(fila["nombre"])
        activo.append(fila["activo"])
        passw.append(fila["pass"])
        id_area_res.append(fila["id_area_res"])
        tc.append(fila["tc"])
        mail.append(fila["mail"])
        id_areat.append(fila["id_areat"])
        id_area_res2.append(fila["id_area_res2"])
        id_area_res3.append(fila["id_area_res3"])
        perm_fsm.append(fila["perm_fsm"])
        tipoPuesto.append(fila["tipoPuesto"])
    

    exportar_datos = { "id_empleado": id_empleado,
                      "id_funcion": id_funcion, 
                      "id_area": id_area,
                      "app": app,
                      "apm": apm,
                      "nombre": nombre,
                      "activo" : activo,
                      "pass": passw,
                      "id_area_res": id_area_res,
                      "tc": tc,
                      "mail": mail,
                      "id_areat": id_areat,
                      "id_area_res2": id_area_res2,
                      "id_area_res3": id_area_res3,
                      "perm_fsm": perm_fsm,
                      "tipoPuesto": tipoPuesto
                      }

    fecha = date.today()
    nombre_archivo = fecha.strftime("%d_%m_%Y")
    df = pd.DataFrame(exportar_datos)
    df.to_excel(f"files/ingresos_{nombre_archivo}.xlsx", index=False)
    print("Archivo listo!")
    
  
formato_datos = generar_formato_tabla()
if sin_datos != []:
    print(f"Faltan datos: --- {sin_datos} ---")
   
 
print("Datos correctos... Generando datos.")
generar_tabla_excel(formato_datos)


